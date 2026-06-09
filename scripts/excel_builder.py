"""
Excel Output Builder
Takes MUS results and writes the formatted workpaper Excel file.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# ── Colours ───────────────────────────────────────────────────────────────────
BLUE_FILL    = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")   # header dark
LIGHT_FILL   = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")   # header light
IS_FILL      = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")   # orange tint IS
MUS_FILL     = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")   # green tint MUS
GREY_FILL    = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")   # alternating row
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
YELLOW_FILL  = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")   # input cells

WHITE_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
NORMAL_FONT  = Font(name="Arial", size=10)
LABEL_FONT   = Font(name="Arial", bold=True, size=10, color="1F4E79")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_cell(ws, cell_ref, value):
    c = ws[cell_ref]
    c.value = value
    c.font = WHITE_FONT
    c.fill = BLUE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER


def _label_cell(ws, cell_ref, value):
    c = ws[cell_ref]
    c.value = value
    c.font = LABEL_FONT
    c.alignment = Alignment(vertical="center")


def _value_cell(ws, cell_ref, value, number_format=None, fill=None, bold=False):
    c = ws[cell_ref]
    c.value = value
    c.font = Font(name="Arial", bold=bold, size=10)
    c.alignment = Alignment(vertical="center")
    c.border = BORDER
    if number_format:
        c.number_format = number_format
    if fill:
        c.fill = fill


def _auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def build_cover_tab(wb, pos_workings, neg_workings, pm, risk_factor, filename, ref_col, value_col):
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "Monetary Unit Sampling – Audit Workpaper"
    title.font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    subtitle = ws["A2"]
    subtitle.value = f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}"
    subtitle.font = Font(name="Arial", size=10, color="808080")
    subtitle.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Section headers helper
    def section(row, label):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws[f"A{row}"]
        c.value = label
        c.font = WHITE_FONT
        c.fill = BLUE_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18

    def row_pair(row, label, value, fmt=None):
        ws[f"A{row}"].value = label
        ws[f"A{row}"].font = NORMAL_FONT
        ws[f"A{row}"].fill = LIGHT_FILL
        ws[f"A{row}"].alignment = Alignment(vertical="center", indent=1)
        ws[f"A{row}"].border = BORDER
        ws.merge_cells(f"A{row}:C{row}")

        c = ws[f"D{row}"]
        c.value = value
        c.font = BOLD_FONT
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = BORDER
        if fmt:
            c.number_format = fmt
        ws.row_dimensions[row].height = 16

    # Inputs section
    section(4, "Sampling Inputs")
    row_pair(5, "Source file", filename)
    row_pair(6, "Reference column", ref_col)
    row_pair(7, "Value column", value_col)
    row_pair(8, "Performance materiality", pm, '#,##0.00')
    row_pair(9, "Risk factor", risk_factor, '0.0000')
    row_pair(10, "Random seed (fixed)", 42)
    row_pair(11, "Sample size rounding", "Always round up (ceiling)")

    # Positive population section
    if pos_workings:
        w = pos_workings
        section(13, "Positive Population Summary")
        row_pair(14, "Total items in population", w["total_count"])
        row_pair(15, "Total population value", w["total_value"], '#,##0.00')
        row_pair(16, "Individually significant items (>= PM) – count", w["is_count"])
        row_pair(17, "Individually significant items (>= PM) – value", w["is_value"], '#,##0.00')
        row_pair(18, "Residual population – count", w["remaining_count"])
        row_pair(19, "Residual population – value", w["remaining_total"], '#,##0.00')
        row_pair(20, "Sample size formula  (Residual Total × Risk Factor ÷ PM)", "=D19*D9/D8", '#,##0.0000')
        row_pair(21, "Raw sample size (n)  [same as above, shown for clarity]", "=D19*D9/D8", '0.0000')
        row_pair(22, "Sample size (rounded up)", w["sample_size"])
        row_pair(23, "Sampling interval", w["interval"], '#,##0.00')
        row_pair(24, "Random start (seed 42)", w["random_start"], '#,##0.00')
        row_pair(25, "MUS items selected", w["mus_selected_count"])
        row_pair(26, "Total items selected (IS + MUS)", w["total_selected"])

    # Negative population section
    if neg_workings:
        w = neg_workings
        section(28, "Negative Population Summary")
        row_pair(29, "Total items in population", w["total_count"])
        row_pair(30, "Total population value (absolute)", w["total_value"], '#,##0.00')
        row_pair(31, "Individually significant items (|value| >= PM) – count", w["is_count"])
        row_pair(32, "Individually significant items (|value| >= PM) – value", w["is_value"], '#,##0.00')
        row_pair(33, "Residual population – count", w["remaining_count"])
        row_pair(34, "Residual population – value", w["remaining_total"], '#,##0.00')
        row_pair(35, "Sample size formula  (Residual Total × Risk Factor ÷ PM)", "=D34*D9/D8", '#,##0.0000')
        row_pair(36, "Raw sample size (n)  [same as above, shown for clarity]", "=D34*D9/D8", '0.0000')
        row_pair(37, "Sample size (rounded up)", w["sample_size"])
        row_pair(38, "Sampling interval", w["interval"], '#,##0.00')
        row_pair(39, "Random start (seed 42)", w["random_start"], '#,##0.00')
        row_pair(40, "MUS items selected", w["mus_selected_count"])
        row_pair(41, "Total items selected (IS + MUS)", w["total_selected"])

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 20


def build_population_tab(wb, tab_name, annotated, original_headers, ref_col, value_col, workings):
    """
    Builds a population tab.
    annotated: list of dicts with keys: transaction, status, running_total
    original_headers: list of column names in original order
    """
    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False

    if not annotated:
        ws["A1"] = "No transactions in this population."
        return

    # Build header row: original columns + new columns
    extra_headers = ["Selection Status", "Running Total (Residual Pop.)"]
    all_headers = original_headers + extra_headers

    for col_idx, h in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = WHITE_FONT
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30

    # Legend below headers
    legend_row = 2
    ws.merge_cells(f"A{legend_row}:{get_column_letter(len(all_headers))}{legend_row}")
    legend_cell = ws[f"A{legend_row}"]
    legend_cell.value = (
        "IS = Individually Significant (|value| >= Performance Materiality, 100% selected)   "
        "MUS = Selected by Monetary Unit Sampling   "
        "Running total applies to residual population only (IS items excluded from MUS walk)"
    )
    legend_cell.font = Font(name="Arial", italic=True, size=9, color="595959")
    legend_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[legend_row].height = 28

    # Data rows start at row 3
    data_start_row = 3

    # We need to track which Excel rows are residual (non-IS) to build running total formulas
    # Running total column index
    status_col_idx = len(original_headers) + 1
    running_col_idx = len(original_headers) + 2
    value_col_idx = original_headers.index(value_col) + 1  # 1-based

    running_col_letter = get_column_letter(running_col_idx)
    value_col_letter = get_column_letter(value_col_idx)

    for row_offset, item in enumerate(annotated):
        excel_row = data_start_row + row_offset
        t = item["transaction"]
        status = item["status"]

        # Choose row fill
        if status == "IS":
            row_fill = IS_FILL
        elif status == "MUS":
            row_fill = MUS_FILL
        elif row_offset % 2 == 0:
            row_fill = WHITE_FILL
        else:
            row_fill = GREY_FILL

        # Write original columns
        for col_idx, h in enumerate(original_headers, start=1):
            val = t.original_row.get(h, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.fill = row_fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")

            # Format value column as number
            if h == value_col:
                try:
                    cell.value = float(str(val).replace("£","").replace("$","").replace(",","").replace(" ",""))
                    cell.number_format = '#,##0.00'
                except (ValueError, TypeError):
                    pass

        # Status column
        status_cell = ws.cell(row=excel_row, column=status_col_idx)
        if status == "IS":
            status_cell.value = "IS"
        elif status == "MUS":
            status_cell.value = "MUS"
        else:
            status_cell.value = ""
        status_cell.font = Font(name="Arial", bold=True, size=10)
        status_cell.fill = row_fill
        status_cell.border = BORDER
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Running total column
        rt_cell = ws.cell(row=excel_row, column=running_col_idx)
        rt_cell.border = BORDER
        rt_cell.fill = row_fill
        rt_cell.alignment = Alignment(horizontal="right", vertical="center")
        rt_cell.font = NORMAL_FONT

        if status == "IS":
            rt_cell.value = "N/A - IS"
            rt_cell.font = Font(name="Arial", italic=True, size=10, color="808080")
        else:
            # Formula: sum of ABS of all non-IS value cells from data_start_row up to this row
            # We use SUMIF on the status column: sum values where status column <> "IS"
            # Simpler: use running sum formula referencing previous running total cell + this ABS value
            if excel_row == data_start_row:
                # First residual row -- find whether there are IS rows above
                # Use a SUMPRODUCT to sum ABS values of non-IS rows up to and including this row
                rt_cell.value = (
                    f'=SUMPRODUCT((${get_column_letter(status_col_idx)}${data_start_row}:'
                    f'${get_column_letter(status_col_idx)}{excel_row}<>"IS")*'
                    f'ABS(${value_col_letter}${data_start_row}:${value_col_letter}{excel_row}))'
                )
            else:
                rt_cell.value = (
                    f'=SUMPRODUCT((${get_column_letter(status_col_idx)}${data_start_row}:'
                    f'${get_column_letter(status_col_idx)}{excel_row}<>"IS")*'
                    f'ABS(${value_col_letter}${data_start_row}:${value_col_letter}{excel_row}))'
                )
            rt_cell.number_format = '#,##0.00'

    # Freeze panes below header + legend
    ws.freeze_panes = f"A{data_start_row}"

    # Auto width
    _auto_width(ws)
    # Force wider on running total col
    ws.column_dimensions[running_col_letter].width = 30


def build_sample_tab(wb, tab_name, annotated, original_headers, ref_col, value_col):
    """Builds a tab showing only selected items (IS and MUS)."""
    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False

    selected = [item for item in annotated if item["status"] in ("IS", "MUS")]

    if not selected:
        ws["A1"] = "No items selected in this population."
        return

    all_headers = original_headers + ["Selection Status"]

    for col_idx, h in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = WHITE_FONT
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30

    value_col_idx = original_headers.index(value_col) + 1
    status_col_idx = len(original_headers) + 1

    for row_offset, item in enumerate(selected):
        excel_row = 2 + row_offset
        t = item["transaction"]
        status = item["status"]
        row_fill = IS_FILL if status == "IS" else MUS_FILL

        for col_idx, h in enumerate(original_headers, start=1):
            val = t.original_row.get(h, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.fill = row_fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            if h == value_col:
                try:
                    cell.value = float(str(val).replace("\xa3","").replace("$","").replace(",","").replace(" ",""))
                    cell.number_format = '#,##0.00'
                except (ValueError, TypeError):
                    pass

        sc = ws.cell(row=excel_row, column=status_col_idx)
        sc.value = status
        sc.font = Font(name="Arial", bold=True, size=10)
        sc.fill = row_fill
        sc.border = BORDER
        sc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[excel_row].height = 16

    # Totals row
    total_row = 2 + len(selected)
    label_col = max(1, value_col_idx - 1)
    lc = ws.cell(row=total_row, column=label_col)
    lc.value = "TOTAL SELECTED"
    lc.font = BOLD_FONT
    lc.fill = LIGHT_FILL
    lc.border = BORDER

    val_col_letter = get_column_letter(value_col_idx)
    tc = ws.cell(row=total_row, column=value_col_idx)
    tc.value = f"=SUM({val_col_letter}2:{val_col_letter}{total_row - 1})"
    tc.font = BOLD_FONT
    tc.fill = LIGHT_FILL
    tc.border = BORDER
    tc.number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    _auto_width(ws)


def build_workbook(
    output_path,
    pos_results,
    neg_results,
    original_headers,
    ref_col,
    value_col,
    pm,
    risk_factor,
    filename,
):
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    pos_workings = pos_results["workings"] if pos_results else None
    neg_workings = neg_results["workings"] if neg_results else None

    # Tab 1: Cover
    build_cover_tab(wb, pos_workings, neg_workings, pm, risk_factor, filename, ref_col, value_col)

    # Tab 2: Positive population
    if pos_results and pos_results["annotated"]:
        build_population_tab(
            wb, "Positive Population",
            pos_results["annotated"],
            original_headers, ref_col, value_col,
            pos_workings,
        )
    else:
        ws = wb.create_sheet("Positive Population")
        ws["A1"] = "No positive transactions in population."

    # Tab 3: Negative population
    if neg_results and neg_results["annotated"]:
        build_population_tab(
            wb, "Negative Population",
            neg_results["annotated"],
            original_headers, ref_col, value_col,
            neg_workings,
        )
    else:
        ws = wb.create_sheet("Negative Population")
        ws["A1"] = "No negative transactions in population."

    # Tab 4: Positive sample (selected only)
    if pos_results and pos_results["annotated"]:
        build_sample_tab(
            wb, "Positive Sample",
            pos_results["annotated"],
            original_headers, ref_col, value_col,
        )
    else:
        ws = wb.create_sheet("Positive Sample")
        ws["A1"] = "No positive transactions selected."

    # Tab 5: Negative sample (selected only)
    if neg_results and neg_results["annotated"]:
        build_sample_tab(
            wb, "Negative Sample",
            neg_results["annotated"],
            original_headers, ref_col, value_col,
        )
    else:
        ws = wb.create_sheet("Negative Sample")
        ws["A1"] = "No negative transactions selected."

    wb.save(output_path)
    print(f"Workbook saved: {output_path}")
