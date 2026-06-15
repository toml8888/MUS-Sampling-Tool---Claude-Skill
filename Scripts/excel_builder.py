"""
Excel Output Builder - New Presentation Format
Tabs: Summary, Population, Sampling form positive, Sampling form negative,
      Positive samples, Negative samples.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from mus_engine import clean_value


# ── Styles ────────────────────────────────────────────────────────────────────
DARK_BLUE    = "1F4E79"
HEADER_FILL  = PatternFill("solid", start_color=DARK_BLUE, end_color=DARK_BLUE)
ACCENT_FILL  = PatternFill("solid", start_color=DARK_BLUE, end_color=DARK_BLUE)  # row 5 accent
INPUT_FILL   = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
IS_FILL      = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
MUS_FILL     = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
GREY_FILL    = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
ABOVE_PM_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
NO_FILL      = PatternFill("none")

WHITE_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
NORMAL_FONT  = Font(name="Arial", size=10)
TITLE_FONT   = Font(name="Arial", bold=True, size=14, color=DARK_BLUE)
INPUT_FONT   = Font(name="Arial", size=10, color="0070C0")
ITALIC_FONT  = Font(name="Arial", italic=True, size=9, color="595959")

med  = Side(style="medium", color="000000")
thin = Side(style="thin",   color="BFBFBF")
BORDER_ALL   = Border(left=thin, right=thin, top=thin, bottom=thin)
MED_ALL      = Border(left=med,  right=med,  top=med,  bottom=med)
MED_LB       = Border(left=med,  bottom=med)
MED_L        = Border(left=med)
MED_LR       = Border(left=med,  right=med)
MED_LRT      = Border(left=med,  right=med,  top=med)
MED_LRB      = Border(left=med,  right=med,  bottom=med)
MED_RT       = Border(right=med, top=med)
MED_RB       = Border(right=med, bottom=med)
MED_R        = Border(right=med)
MED_T        = Border(top=med)
MED_B        = Border(bottom=med)
THIN_B       = Border(bottom=thin)
MED_R_THIN_B = Border(right=med, bottom=thin)
THIN_B_MED_R = Border(bottom=thin, right=med)


# ── Helpers ───────────────────────────────────────────────────────────────────
def _auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _cell(ws, coord, value=None, font=None, fill=None, alignment=None,
          number_format=None, border=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if alignment:
        c.alignment = alignment
    if number_format:
        c.number_format = number_format
    if border:
        c.border = border
    return c


def _apply_outer_border(ws, min_row, max_row, min_col, max_col):
    """Apply medium border around a rectangular range."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            left   = med if c == min_col else cell.border.left
            right  = med if c == max_col else cell.border.right
            top    = med if r == min_row else cell.border.top
            bottom = med if r == max_row else cell.border.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def _add_image(ws, img_path, anchor):
    if os.path.exists(img_path):
        img = XLImage(img_path)
        img.anchor = anchor
        ws.add_image(img)


# ── Tab 1: Summary ────────────────────────────────────────────────────────────
def build_summary_tab(wb, pm, risk_factor, class_col):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    _cell(ws, "B2", "PM", font=BOLD_FONT)
    c = ws["C2"]
    c.value = pm
    c.font = INPUT_FONT
    c.number_format = '#,##0.00'

    _cell(ws, "B3", "Total positive", font=NORMAL_FONT)
    _cell(ws, "C3", f'=SUMIF(Population!{class_col}:{class_col},"Positive",Population!B:B)',
          font=NORMAL_FONT, number_format='#,##0.00')

    _cell(ws, "B4", "Total negative", font=NORMAL_FONT)
    _cell(ws, "C4", f'=SUMIF(Population!{class_col}:{class_col},"Negative",Population!B:B)',
          font=NORMAL_FONT, number_format='#,##0.00')

    _cell(ws, "B5", "Total population", font=NORMAL_FONT)
    _cell(ws, "C5", '=C3+C4', font=NORMAL_FONT, number_format='#,##0.00')

    _cell(ws, "B7", "Risk factor", font=BOLD_FONT)
    c = ws["C7"]
    c.value = risk_factor
    c.font = INPUT_FONT
    c.number_format = '0.0000'

    _cell(ws, "B8", "Total samples", font=BOLD_FONT)
    _cell(ws, "C8",
          "=IFERROR('Sampling form positive'!E37,0)+IFERROR('Sampling form negative'!E37,0)",
          font=NORMAL_FONT, number_format='0')

    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18


# ── Tab 2: Population ─────────────────────────────────────────────────────────
def build_population_tab(wb, all_rows, identifier_col, value_col,
                         original_headers, decimal_system):
    """
    A  = identifier (population number column)
    B  = value (numeric)
    C onwards = remaining columns (up to 15, excluding identifier and value)
    Then Above PM, then Classification
    Returns (above_pm_col_letter, classification_col_letter)
    """
    ws = wb.create_sheet("Population")
    ws.sheet_view.showGridLines = False

    # Exclude BOTH identifier_col and value_col so neither appears twice
    other_headers = [h for h in original_headers
                     if h != identifier_col and h != value_col][:15]
    display_headers = [identifier_col, value_col] + other_headers

    # Data column headers (A onwards)
    for col_idx, h in enumerate(display_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = WHITE_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_ALL
    ws.row_dimensions[1].height = 30

    # Above PM header
    above_col = len(display_headers) + 1
    c = ws.cell(row=1, column=above_col, value="Above PM")
    c.font = Font(name="Arial", bold=True, size=10)
    c.fill = ABOVE_PM_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER_ALL

    # Classification header
    class_col = len(display_headers) + 2
    c = ws.cell(row=1, column=class_col, value="Classification")
    c.font = Font(name="Arial", bold=True, size=10)
    c.fill = ABOVE_PM_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER_ALL

    above_col_letter = get_column_letter(above_col)
    class_col_letter = get_column_letter(class_col)

    # Data rows
    for row_offset, row_data in enumerate(all_rows):
        excel_row = 2 + row_offset

        # Col A: identifier
        c = ws.cell(row=excel_row, column=1, value=row_data.get(identifier_col, ""))
        c.font = NORMAL_FONT
        c.border = BORDER_ALL
        c.alignment = Alignment(vertical="center")

        # Col B: value (numeric)
        raw_val = row_data.get(value_col, "")
        parsed = clean_value(raw_val, decimal_system)
        c = ws.cell(row=excel_row, column=2, value=parsed if parsed is not None else raw_val)
        c.font = NORMAL_FONT
        c.border = BORDER_ALL
        c.alignment = Alignment(horizontal="right", vertical="center")
        if parsed is not None:
            c.number_format = '#,##0.00'

        # Cols C onwards: other columns
        for col_offset, h in enumerate(other_headers):
            c = ws.cell(row=excel_row, column=3 + col_offset, value=row_data.get(h, ""))
            c.font = NORMAL_FONT
            c.border = BORDER_ALL
            c.alignment = Alignment(vertical="center")

        # Above PM formula
        m = ws.cell(row=excel_row, column=above_col,
                    value=f'=IF(ABS(B{excel_row})>Summary!$C$2,"Yes","")')
        m.font = NORMAL_FONT
        m.border = BORDER_ALL
        m.alignment = Alignment(horizontal="center", vertical="center")

        # Classification formula
        n = ws.cell(row=excel_row, column=class_col,
                    value=f'=IF(B{excel_row}>0,"Positive",IF(B{excel_row}<0,"Negative","Zero"))')
        n.font = NORMAL_FONT
        n.border = BORDER_ALL
        n.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    _auto_width(ws)
    ws.column_dimensions[above_col_letter].width = 12
    ws.column_dimensions[class_col_letter].width = 14

    return above_col_letter, class_col_letter


# ── Tab 3 & 4: Sampling Form ──────────────────────────────────────────────────
# Column layout matches example:
#  B=label area (merged B:D), E F G H I J = risk factor columns, K = right border
#  Rows: 5=blue accent, 8=blue title bar (B:K merged), then form body down to ~39

def build_sampling_form_tab(wb, tab_name, population_sign, img1_path,
                            above_pm_col, class_col):
    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False

    sign = population_sign  # "Positive" or "Negative"

    # ── Column widths (matching example) ──────────────────────────────────────
    ws.column_dimensions["B"].width = 12.22
    ws.column_dimensions["C"].width = 30.22
    ws.column_dimensions["D"].width = 17.55
    ws.column_dimensions["E"].width = 16.11
    ws.column_dimensions["F"].width = 16.33
    ws.column_dimensions["G"].width = 16.22
    ws.column_dimensions["H"].width = 16.22
    ws.column_dimensions["I"].width = 16.22
    ws.column_dimensions["J"].width = 19.78

    # ── Row 5: blue accent line ────────────────────────────────────────────────
    for col in range(2, 11):  # B to J
        c = ws.cell(row=5, column=col)
        c.fill = ACCENT_FILL
    ws.row_dimensions[5].height = 6

    # ── Row 8: "Sampling Form" title bar ──────────────────────────────────────
    ws.merge_cells("B8:J8")
    c = ws["B8"]
    c.value = f"Sampling Form \u2013 {sign} Population"
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = ACCENT_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    # Medium border all around B8:J8
    _apply_outer_border(ws, 8, 8, 2, 10)
    ws.row_dimensions[8].height = 20

    # ── Outer border: medium left on B, medium right on J, rows 8-39 ──────────
    LAST_ROW = 39
    for row in range(8, LAST_ROW + 1):
        # Left border col B
        c = ws.cell(row=row, column=2)
        existing = c.border
        c.border = Border(left=med, right=existing.right,
                          top=existing.top, bottom=existing.bottom)
        # Right border col J
        c = ws.cell(row=row, column=10)
        existing = c.border
        c.border = Border(left=existing.left, right=med,
                          top=existing.top, bottom=existing.bottom)

    # ── Row 9: top of inner area (medium top) ─────────────────────────────────
    for col in range(2, 11):
        c = ws.cell(row=9, column=col)
        existing = c.border
        c.border = Border(left=existing.left, right=existing.right,
                          top=med, bottom=existing.bottom)

    # ── Bottom of form: medium bottom row 39 ─────────────────────────────────
    for col in range(2, 11):
        c = ws.cell(row=LAST_ROW, column=col)
        existing = c.border
        c.border = Border(left=existing.left, right=existing.right,
                          top=existing.top, bottom=med)

    # ── Key row ───────────────────────────────────────────────────────────────
    _cell(ws, "B10", "Key:", font=BOLD_FONT)
    c = ws["C10"]
    c.value = "Cells requiring input"
    c.font = INPUT_FONT
    c.fill = INPUT_FILL
    c.border = MED_ALL
    c.alignment = Alignment(horizontal="center")

    # ── Random seed row ───────────────────────────────────────────────────────
    ws.merge_cells("B12:D12")
    _cell(ws, "B12", "Random seed (fixed):", font=NORMAL_FONT,
          alignment=Alignment(vertical="center"))
    c = ws["E12"]
    c.value = 42
    c.font = NORMAL_FONT
    c.border = MED_ALL
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.number_format = '0'

    # ── Guidance text ─────────────────────────────────────────────────────────
    ws.merge_cells("B14:J14")
    c = ws["B14"]
    c.value = ("If you are testing multiple assertions, use the highest risk level "
               "identified for the assertions being tested.")
    c.font = ITALIC_FONT
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[14].height = 30

    ws.merge_cells("B15:J15")
    c = ws["B15"]
    c.value = ("Ensure that the sample provides a reasonable basis on which to draw "
               "conclusions about the population from which the sample is selected.")
    c.font = ITALIC_FONT
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[15].height = 30

    # ── "Overall Risk Level" label ────────────────────────────────────────────
    ws.merge_cells("G17:J17")
    c = ws["G17"]
    c.value = "Overall Risk Level for the Assertion"
    c.font = BOLD_FONT
    c.alignment = Alignment(horizontal="center")
    c.border = Border(right=med, bottom=thin)

    # ── Risk factor table ─────────────────────────────────────────────────────
    # Header row 19: B:D = label, E F G H I J = risk levels
    ws.merge_cells("B19:D19")
    c = ws["B19"]
    c.value = "Table of Risk Factors"
    c.font = BOLD_FONT
    c.border = Border(left=med, bottom=thin)

    risk_headers = [
        (5, "E", "1 (No RMM)"),
        (6, "F", "2 (Low RMM)"),
        (7, "G", "3 (Moderate RMM)"),
        (8, "H", "4 (Elevated RMM)"),
        (9, "I", "5 (Significant RMM)"),
        (10, "J", ""),  # right border only
    ]
    for col_idx, col_letter, label in risk_headers:
        c = ws.cell(row=19, column=col_idx, value=label if label else None)
        c.font = BOLD_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=False)
        left_b  = med if col_idx == 10 else None
        right_b = med if col_idx == 10 else None
        c.border = Border(
            left=left_b, right=right_b,
            bottom=thin
        )

    # Three risk rows: 20-21 (row 1), 22-23 (row 2), 24-25 (row 3)
    rf_rows = [
        (20, "Tests of details, AND analytical procedures AND tests of control performed",
         [0.2, 0.4, 0.6, 0.8, 1.0]),
        (22, "Tests of details, and analytical procedures OR tests of control performed",
         [0.4, 0.8, 1.2, 1.6, 2.0]),
        (24, "Tests of details only performed",
         [0.6, 1.1, 1.6, 2.3, 3.0]),
    ]

    for start_row, label, values in rf_rows:
        is_last = (start_row == 24)
        # Merge label across B:D for two rows
        ws.merge_cells(f"B{start_row}:D{start_row + 1}")
        c = ws[f"B{start_row}"]
        c.value = label
        c.font = NORMAL_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center")
        bottom_b = thin if is_last else None
        c.border = Border(left=med, bottom=bottom_b if is_last else None)

        # Merge each value column for two rows, E-I
        for col_offset, val in enumerate(values):
            col_idx = 5 + col_offset  # E=5 through I=9
            ws.merge_cells(f"{get_column_letter(col_idx)}{start_row}:{get_column_letter(col_idx)}{start_row+1}")
            c = ws.cell(row=start_row, column=col_idx, value=val)
            c.font = NORMAL_FONT
            c.number_format = '0.0'
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(bottom=thin if is_last else None)

        # Right border col J for both rows
        for r in range(start_row, start_row + 2):
            c = ws.cell(row=r, column=10)
            existing = c.border
            c.border = Border(left=existing.left, right=med,
                              top=existing.top,
                              bottom=thin if (is_last and r == start_row + 1) else existing.bottom)

    # above_pm_col / class_col are the actual column letters in the Population tab
    apm = above_pm_col   # e.g. "M"
    cls = class_col      # e.g. "N"

    # ── Sampling inputs section ───────────────────────────────────────────────
    input_rows = [
        (27, "Sampling risk factor taken from the above table:", "=Summary!C7",     "0.0000", True),
        (29, "Performance materiality (PM)",                      "=Summary!C2",     "#,##0.00", True),
        (31, f"Total monetary value of {sign.lower()} population",
             f'=SUMIF(Population!{cls}:{cls},"{sign}",Population!B:B)', "#,##0.00", False),
        (32, f"Value of items above PM ({sign.lower()})",
             f'=SUMIFS(Population!B:B,Population!{apm}:{apm},"Yes",Population!{cls}:{cls},"{sign}")', "#,##0.00", False),
        (33, f"Number of items above PM ({sign.lower()})",
             f'=COUNTIFS(Population!{apm}:{apm},"Yes",Population!{cls}:{cls},"{sign}")', "0", False),
        (34, "Value of residual population",
             f'=SUMIFS(Population!B:B,Population!{apm}:{apm},"",Population!{cls}:{cls},"{sign}")', "#,##0.00", False),
    ]

    for row_num, label, formula, num_fmt, is_input in input_rows:
        ws.merge_cells(f"B{row_num}:D{row_num}")
        lc = ws[f"B{row_num}"]
        lc.value = label
        lc.font = NORMAL_FONT
        lc.alignment = Alignment(vertical="center")
        lc.border = Border(left=med)

        ec = ws[f"E{row_num}"]
        ec.value = formula
        ec.font = INPUT_FONT if is_input else NORMAL_FONT
        ec.fill = INPUT_FILL if is_input else NO_FILL
        ec.border = MED_ALL if is_input else BORDER_ALL
        ec.alignment = Alignment(horizontal="right", vertical="center")
        ec.number_format = num_fmt

        # Right border on J for these rows
        c = ws.cell(row=row_num, column=10)
        existing = c.border
        c.border = Border(left=existing.left, right=med,
                          top=existing.top, bottom=existing.bottom)

    # ── Sample size row 37 (no merge across rows) ────────────────────────────
    ws.merge_cells("B37:D37")
    lc = ws["B37"]
    lc.value = "Calculated sample size (residual value \u00d7 risk factor \u00f7 PM)"
    lc.font = BOLD_FONT
    lc.alignment = Alignment(wrap_text=False, vertical="center")
    lc.border = Border(left=med)

    sc = ws["E37"]
    sc.value = "=CEILING(ABS(E34)*E27/E29,1)"
    sc.font = BOLD_FONT
    sc.fill = NO_FILL
    sc.border = MED_ALL
    sc.number_format = '0'
    sc.alignment = Alignment(horizontal="right", vertical="center")

    # Right border on J37
    c = ws.cell(row=37, column=10)
    existing = c.border
    c.border = Border(left=existing.left, right=med,
                      top=existing.top, bottom=existing.bottom)

    # ── Image: logo top-left, smaller size ───────────────────────────────────
    if os.path.exists(img1_path):
        img = XLImage(img1_path)
        img.width  = 120
        img.height = 42
        img.anchor = "B1"
        ws.add_image(img)


# ── Tab 5 & 6: Sample tabs ────────────────────────────────────────────────────
def build_sample_tab(wb, tab_name, annotated, identifier_col, value_col,
                     original_headers, decimal_system, population_sign):
    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False

    selected = [item for item in annotated if item["status"] in ("IS", "MUS")]

    if not selected:
        ws["A1"] = f"No {population_sign.lower()} items selected."
        ws["A1"].font = NORMAL_FONT
        return

    other_headers = [h for h in original_headers
                     if h != identifier_col and h != value_col][:15]
    display_headers = [identifier_col, value_col] + other_headers + ["Selection Status"]

    for col_idx, h in enumerate(display_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = WHITE_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_ALL
    ws.row_dimensions[1].height = 30

    status_col_idx = len(display_headers)

    for row_offset, item in enumerate(selected):
        excel_row = 2 + row_offset
        t = item["transaction"]
        status = item["status"]
        row_fill = IS_FILL if status == "IS" else MUS_FILL

        # Col A: identifier
        c = ws.cell(row=excel_row, column=1, value=t.original_row.get(identifier_col, ""))
        c.font = NORMAL_FONT
        c.fill = row_fill
        c.border = BORDER_ALL
        c.alignment = Alignment(vertical="center")

        # Col B: value
        raw_val = t.original_row.get(value_col, "")
        parsed = clean_value(raw_val, decimal_system)
        c = ws.cell(row=excel_row, column=2, value=parsed if parsed is not None else raw_val)
        c.font = NORMAL_FONT
        c.fill = row_fill
        c.border = BORDER_ALL
        c.alignment = Alignment(horizontal="right", vertical="center")
        if parsed is not None:
            c.number_format = '#,##0.00'

        # Other cols
        for col_offset, h in enumerate(other_headers):
            c = ws.cell(row=excel_row, column=3 + col_offset, value=t.original_row.get(h, ""))
            c.font = NORMAL_FONT
            c.fill = row_fill
            c.border = BORDER_ALL
            c.alignment = Alignment(vertical="center")

        # Selection Status
        sc = ws.cell(row=excel_row, column=status_col_idx, value=status)
        sc.font = Font(name="Arial", bold=True, size=10)
        sc.fill = row_fill
        sc.border = BORDER_ALL
        sc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[excel_row].height = 16

    # Total row
    total_row = 2 + len(selected)
    lc = ws.cell(row=total_row, column=1, value="TOTAL SELECTED")
    lc.font = BOLD_FONT
    lc.border = BORDER_ALL

    tc = ws.cell(row=total_row, column=2,
                 value=f"=SUM(B2:B{total_row - 1})")
    tc.font = BOLD_FONT
    tc.border = BORDER_ALL
    tc.number_format = '#,##0.00'
    tc.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A2"
    _auto_width(ws)


# ── Main entry ────────────────────────────────────────────────────────────────
def build_workbook(
    output_path,
    pos_results,
    neg_results,
    original_headers,
    popnum_col,
    value_col,
    pm,
    risk_factor,
    filename,
    decimal_system,
    reconciliation,
    exclusions,
):
    # Locate images relative to this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    img1_path = os.path.join(script_dir, "image1.png")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Collect all rows preserving original order
    all_rows = []
    if pos_results and pos_results.get("annotated"):
        all_rows += [item["transaction"].original_row for item in pos_results["annotated"]]
    if neg_results and neg_results.get("annotated"):
        all_rows += [item["transaction"].original_row for item in neg_results["annotated"]]

    identifier_col = popnum_col

    # Tab 2: Population first -- need col letters before building Summary and sampling forms
    above_pm_col, class_col = build_population_tab(
        wb, all_rows, identifier_col, value_col, original_headers, decimal_system)

    # Tab 1: Summary (insert at position 0 so it appears first)
    build_summary_tab(wb, pm, risk_factor, class_col)
    # Move Summary to front
    wb.move_sheet("Summary", offset=-wb.sheetnames.index("Summary"))

    # Tab 3: Sampling form positive
    build_sampling_form_tab(wb, "Sampling form positive", "Positive",
                            img1_path, above_pm_col, class_col)

    # Tab 4: Sampling form negative
    build_sampling_form_tab(wb, "Sampling form negative", "Negative",
                            img1_path, above_pm_col, class_col)

    # Tab 5: Positive samples
    if pos_results and pos_results.get("annotated"):
        build_sample_tab(wb, "Positive samples", pos_results["annotated"],
                         identifier_col, value_col, original_headers, decimal_system, "Positive")
    else:
        ws = wb.create_sheet("Positive samples")
        ws["A1"] = "No positive items selected."

    # Tab 6: Negative samples
    if neg_results and neg_results.get("annotated"):
        build_sample_tab(wb, "Negative samples", neg_results["annotated"],
                         identifier_col, value_col, original_headers, decimal_system, "Negative")
    else:
        ws = wb.create_sheet("Negative samples")
        ws["A1"] = "No negative items selected."

    wb.save(output_path)
    print(f"Workbook saved: {output_path}")
